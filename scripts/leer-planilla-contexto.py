"""
Vuelca un rango de filas de una hoja de la planilla, con su encabezado.

    python scripts/leer-planilla-contexto.py Contexto 110 135

Existe porque las marcas de tecnica de la planilla aparecen en tres lugares
distintos y no siempre coinciden — sin ver la fila entera no se sabe cual es la
vigente.
"""
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

RUTA = Path("data/Rutina gym - Claude (TEST CELULAR).xlsx")


def main() -> int:
    hoja = sys.argv[1] if len(sys.argv) > 1 else "Contexto"
    desde = int(sys.argv[2]) if len(sys.argv) > 2 else 1
    hasta = int(sys.argv[3]) if len(sys.argv) > 3 else desde + 25

    wb = load_workbook(RUTA, data_only=True)
    ws = wb[hoja]
    for i, fila in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < desde or i > hasta:
            continue
        celdas = ["" if c is None else str(c).strip() for c in fila]
        if not any(celdas):
            continue
        print(f"{i:4}: " + " | ".join(c for c in celdas if c))
    return 0


if __name__ == "__main__":
    sys.exit(main())
