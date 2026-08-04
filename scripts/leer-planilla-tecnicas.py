"""
Busca en la planilla original que ejercicios llevan dropset u otra tecnica.

    python scripts/leer-planilla-tecnicas.py

La planilla es la fuente: la app tiene que reflejarla, no al reves. Se listan
las hojas, sus encabezados y toda celda que mencione una tecnica, con su fila
completa — sin eso no se sabe A QUE ejercicio corresponde la marca.
"""
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

RUTA = Path("data/Rutina gym - Claude (TEST CELULAR).xlsx")
PATRON = re.compile(
    r"drop\s*-?\s*set|\bds\b|rest\s*-?\s*pause|myo|cluster|parcial|excentric|isometr|descendente",
    re.I,
)


def main() -> int:
    if not RUTA.exists():
        print(f"no esta {RUTA}")
        return 1
    wb = load_workbook(RUTA, data_only=True)
    print(f"hojas: {wb.sheetnames}\n")

    for hoja in wb.sheetnames:
        ws = wb[hoja]
        filas = list(ws.iter_rows(values_only=True))
        if not filas:
            continue
        encabezado = [str(c) if c is not None else "" for c in filas[0]]
        hits = []
        for i, fila in enumerate(filas):
            texto = " | ".join("" if c is None else str(c) for c in fila)
            if PATRON.search(texto):
                hits.append((i + 1, fila))
        if not hits:
            continue
        print(f"=== {hoja} ({len(filas)} filas) ===")
        print("  encabezado:", [c for c in encabezado if c][:14])
        for n, fila in hits:
            celdas = ["" if c is None else str(c).strip() for c in fila]
            print(f"  fila {n}: " + " | ".join(c for c in celdas if c))
        print()
    return 0


if __name__ == "__main__":
    sys.exit(main())
